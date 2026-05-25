#!/usr/bin/env python3
"""
最优采购表生成器 v2.0 — 写回 Excel

输入: 原始采购表 Excel + 搜索结果 CSV
输出: 原始采购表（新增：最优单价、发货地、运费、总价 四列）
"""

import sys, csv, xlrd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

SHIPPING = {"上海": 350, "佛山": 150, "湛江": 170, "武汉": 300}

def read_excel_materials(path):
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    materials = []
    for r in range(1, ws.nrows):
        seq = ws.cell_value(r, 0)
        material = str(ws.cell_value(r, 1)).strip()
        spec = str(ws.cell_value(r, 2)).strip()
        demand = ws.cell_value(r, 3)
        if material and spec and demand:
            materials.append({
                "row": r + 1,  # 1-indexed for openpyxl
                "seq": seq,
                "material": material,
                "spec": spec,
                "demand": float(demand),
            })
    return materials

def read_prices(path):
    prices = []
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            prices.append(row)
    return prices

def find_best(material, spec, demand, prices):
    candidates = []
    for p in prices:
        if p["material"].strip() == material and p["spec"].strip() == spec:
            try:
                unit_price = float(p["unit_price"])
                location = p["location"].strip()
                shipping = SHIPPING.get(location, 0)
                total = unit_price * demand + shipping
                candidates.append({
                    "location": location,
                    "unit_price": unit_price,
                    "shipping": shipping,
                    "total": total,
                })
            except (ValueError, KeyError):
                pass
    
    if not candidates:
        return None
    return min(candidates, key=lambda x: x["total"])

def write_to_excel(input_path, materials, results):
    wb = load_workbook(input_path)
    ws = wb.active
    
    # 加表头
    headers = ["最优单价(元/T)", "发货地", "运费(元)", "总价(元)"]
    for i, h in enumerate(headers):
        col = 5 + i  # E, F, G, H
        cell = ws.cell(row=1, column=col, value=h)
        # 复制表头样式
        ref = ws.cell(row=1, column=1)
        if ref.font:
            cell.font = copy(ref.font)
        if ref.fill:
            cell.fill = copy(ref.fill)
    
    for m in materials:
        r = m["row"]
        res = results.get((m["material"], m["spec"]))
        if res:
            ws.cell(row=r, column=5, value=res["unit_price"])
            ws.cell(row=r, column=6, value=res["location"])
            ws.cell(row=r, column=7, value=res["shipping"])
            ws.cell(row=r, column=8, value=round(res["total"], 2))
        else:
            ws.cell(row=r, column=6, value="无报价")
    
    # 自动列宽
    for col in range(1, 9):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = str(ws.cell(row=row, column=col).value or "")
            max_len = max(max_len, len(val.encode("gbk", errors="replace")))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 30)
    
    output = Path(input_path).parent / f"{Path(input_path).stem}_最优采购表.xlsx"
    wb.save(output)
    return output

def main():
    if len(sys.argv) < 3:
        print("用法: python3 fill_table.py <采购表.xls> <搜索结果.csv>")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    prices_path = sys.argv[2]
    
    materials = read_excel_materials(excel_path)
    prices = read_prices(prices_path)
    
    print(f"材料: {len(materials)} 种, 报价: {len(prices)} 条")
    
    results = {}
    total_cost = 0
    print(f"\n{'材质':<15} {'规格':<15} {'需求(T)':>8} {'最优发货地':<8} {'单价':>10} {'运费':>8} {'总价':>12}")
    print("-" * 80)
    
    for m in materials:
        best = find_best(m["material"], m["spec"], m["demand"], prices)
        if best:
            results[(m["material"], m["spec"])] = best
            total_cost += best["total"]
            print(f"{m['material']:<15} {m['spec']:<15} {m['demand']:>8.0f} {best['location']:<8} {best['unit_price']:>10.0f} {best['shipping']:>8.0f} {best['total']:>12.0f}")
        else:
            print(f"{m['material']:<15} {m['spec']:<15} {m['demand']:>8.0f} {'无报价':<8}")
    
    print("-" * 80)
    print(f"{'合计':>56} {total_cost:>12.0f}")
    
    output = write_to_excel(excel_path, materials, results)
    print(f"\n✅ 最优采购表: {output}")

if __name__ == "__main__":
    main()
